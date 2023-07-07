import openpyxl

def getrowcount(file,sheetname):
    book=openpyxl.load_workbook(file)
    sheet=book[sheetname]
    return (sheet.max.row)

def readdata(file,sheetname,rowno,colno):
    book=openpyxl.load_workbook(file)
    sheet=book[sheetname]
    return sheet.cell(row=rowno,column=colno).value

def writedata(file,sheetname,rowno,colno,data):
    book=openpyxl(file)
    sheet=book[sheetname]
    sheet.cell(row=rowno,column=colno).value
    book.save(file)